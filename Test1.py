import openpyxl

# 创建一个Workbook对象

wb = openpyxl.Workbook()

# 创建一个Sheet对象

mySheet = wb.create_sheet(title="Mysheet",index=0)

# 再创建一个Sheet对象

anotherSheet = wb.create_sheet(title='AnotherSheet9990',index=1, )
wb.create_sheet("abc",2)

# 获取活动的sheet

activeSheet = wb.get_active_sheet()

# 设置活动表颜色

activeSheet.sheet_properties.tabColor = "205EB2"

# 设置anotherSheet的标题

anotherSheet.title = "test"

# 选择Cell对象（B4单元格并赋值）

directionCell = activeSheet.cell(row=4, column=2)

directionCell.value = "找到这个单元格"

# 还可以知道单元格的行列最大指

anotherSheet['A1'].value = "activesheet最大行：" + str(activeSheet.max_row)

anotherSheet['A2'].value = "activesheet最大列：" + str(activeSheet.max_column)

# 最后保存workbook

wb.save("test.xlsx")